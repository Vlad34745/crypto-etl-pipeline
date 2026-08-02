"""
Crypto ETL Pipeline & Business Intelligence Dashboard.

Fetches live market data from the CoinGecko API, appends it to a
historical Excel log (with automatic timestamped backups), and
rebuilds a styled "Executive Dashboard" summary sheet.

Run directly:
    python crypto_automation.py
"""

import datetime
import logging
import os
import shutil
import sys
import time

import openpyxl
import pandas as pd
import requests
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import config

logger = logging.getLogger("crypto_etl")


def setup_logging() -> None:
    """Log to both a rotating-friendly file and the console."""
    logging.basicConfig(
        level=getattr(logging, config.LOG_LEVEL.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(config.LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def resolve_coin_ids(user_input: str) -> str:
    """Turn free-typed tickers/ids ("btc eth sol" or "btc,eth,sol") into a
    CoinGecko-ready comma-separated id string ("bitcoin,ethereum,solana").

    Tokens found in config.TICKER_TO_COIN_ID are translated; anything else
    is assumed to already be a valid CoinGecko id and passed through
    lowercased, as-is.
    """
    tokens = [t.strip() for t in user_input.replace(",", " ").split() if t.strip()]
    resolved = []
    for token in tokens:
        coin_id = config.TICKER_TO_COIN_ID.get(token.upper(), token.lower())
        resolved.append(coin_id)
    return ",".join(resolved)


def prompt_for_coins(default_coin_ids: str) -> str:
    """Interactively ask which coins to track; Enter keeps the current default."""
    answer = input(
        f"Які монети відстежувати? (тікери через пробіл, напр. \"btc eth sol\"; "
        f"Enter — залишити поточні [{default_coin_ids}]): "
    ).strip()
    if not answer:
        return default_coin_ids
    resolved = resolve_coin_ids(answer)
    print(f"-> Обрано: {resolved}")
    return resolved


# =============================================================================
# STEP 1: Automated Data Ingestion & API Integration
# =============================================================================
def fetch_crypto_data_with_retry(retries: int = None, delay: int = None):
    """Call the CoinGecko markets endpoint, retrying on 429 / network errors."""
    retries = config.RETRIES if retries is None else retries
    delay = config.RETRY_DELAY_SECONDS if delay is None else delay

    params = {
        "vs_currency": "usd",
        "ids": config.COIN_IDS,
        "order": "market_cap_desc",
        "per_page": 100,
        "page": 1,
        "sparkline": "false",
        "price_change_percentage": "24h",
    }

    for attempt in range(retries):
        try:
            response = requests.get(config.API_URL, params=params, timeout=30)
            if response.status_code == 429:
                logger.warning(
                    "Rate limit hit (429). Retrying in %ss... (Attempt %s/%s)",
                    delay, attempt + 1, retries,
                )
                time.sleep(delay)
                continue
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error("Attempt %s failed: %s", attempt + 1, e)
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                return None
    return None


def clean_raw_data(raw_json) -> pd.DataFrame:
    """Turn the raw CoinGecko JSON payload into a tidy snapshot DataFrame."""
    df_raw = pd.DataFrame(raw_json)
    columns_to_keep = [
        "name", "symbol", "current_price", "market_cap",
        "total_volume", "price_change_percentage_24h",
    ]
    df_clean = df_raw[columns_to_keep].copy()

    # Convert API percentages (e.g. 0.48) into true decimal fractions (0.0048)
    df_clean["price_change_percentage_24h"] = df_clean["price_change_percentage_24h"] / 100.0
    df_clean["snapshot_time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df_clean["symbol"] = df_clean["symbol"].str.upper()
    return df_clean


def validate_data(df: pd.DataFrame, expected_coin_count: int = None) -> None:
    """Sanity-check the fetched snapshot before it touches disk.

    Raises ValueError with a clear message if something looks wrong, so a
    malformed API response never silently corrupts the historical file.
    """
    required_columns = {
        "name", "symbol", "current_price", "market_cap",
        "total_volume", "price_change_percentage_24h", "snapshot_time",
    }
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Missing expected column(s) in fetched data: {sorted(missing)}")

    if df.empty:
        raise ValueError("Fetched data is empty — API may have returned no results.")

    if df["current_price"].isna().any():
        raise ValueError("One or more coins have a null current_price.")

    expected = expected_coin_count or len(config.COIN_IDS.split(","))
    if len(df) != expected:
        logger.warning(
            "Expected %s coins (from COIN_IDS) but received %s rows. "
            "A coin ID may be invalid or delisted.",
            expected, len(df),
        )


# =============================================================================
# STEP 2: Incremental Data Accumulation & History Tracking
# =============================================================================
def backup_existing_file(output_file: str, backup_dir: str, max_backups: int) -> None:
    """Snapshot the current workbook before overwriting it, then prune old backups."""
    if not os.path.exists(output_file):
        return

    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = os.path.join(backup_dir, f"crypto_history_backup_{timestamp_str}.xlsx")
    shutil.copyfile(output_file, backup_filename)
    logger.info("Backup saved to '%s'", backup_filename)

    existing_backups = sorted(
        f for f in os.listdir(backup_dir)
        if f.startswith("crypto_history_backup_") and f.endswith(".xlsx")
    )
    backups_to_delete = existing_backups[:-max_backups] if len(existing_backups) > max_backups else []
    for old_backup in backups_to_delete:
        os.remove(os.path.join(backup_dir, old_backup))
    if backups_to_delete:
        logger.info("Retention: removed %s old backup(s), keeping last %s.",
                     len(backups_to_delete), max_backups)


def load_and_merge_history(output_file: str, df_new: pd.DataFrame) -> pd.DataFrame:
    """Merge the new snapshot into the historical timeline, deduplicated."""
    if os.path.exists(output_file):
        logger.info("Loading historical timeline from '%s'...", output_file)
        df_historical = pd.read_excel(output_file, sheet_name="Crypto Market Timeline")
        df_updated = pd.concat([df_historical, df_new], ignore_index=True)
        df_updated.drop_duplicates(subset=["symbol", "snapshot_time"], keep="first", inplace=True)
    else:
        logger.info("No existing tracking file found. Initializing new dataset.")
        df_updated = df_new.copy()

    df_updated.sort_values(by=["name", "snapshot_time"], inplace=True)
    df_updated.reset_index(drop=True, inplace=True)

    logger.info("Sync complete. Total rows: %s", len(df_updated))
    return df_updated


# =============================================================================
# STEP 3: Executive BI Reporting & Advanced Excel Styling
# =============================================================================
def build_excel_report(df_updated: pd.DataFrame, output_file: str) -> None:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_updated.to_excel(writer, sheet_name="Crypto Market Timeline", index=False)

    wb = openpyxl.load_workbook(output_file)
    ws_timeline = wb["Crypto Market Timeline"]

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws_dash = wb.create_sheet("Dashboard", index=0)

    FONT_NAME = "Segoe UI"
    HEADER_COLOR = "0A5C36"   # Premium Emerald Green
    ACCENT_BG = "E8F5E9"      # Soft Mint Green for KPI blocks

    font_title = Font(name=FONT_NAME, size=18, bold=True, color=HEADER_COLOR)
    font_subtitle = Font(name=FONT_NAME, size=10, italic=True, color="555555")
    font_section = Font(name=FONT_NAME, size=12, bold=True, color="333333")
    font_kpi_label = Font(name=FONT_NAME, size=10, bold=True, color="666666")
    font_kpi_value = Font(name=FONT_NAME, size=14, bold=True, color="111111")
    font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=FONT_NAME, size=11, color="333333")
    font_positive = Font(name=FONT_NAME, size=11, color="0E622F", bold=True)
    font_negative = Font(name=FONT_NAME, size=11, color="9C0006", bold=True)

    fill_header = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    fill_kpi = PatternFill(start_color=ACCENT_BG, end_color=ACCENT_BG, fill_type="solid")
    thin_side = Side(style="thin", color="D3D3D3")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # --- Dashboard header ---
    ws_dash.views.sheetView[0].showGridLines = True
    ws_dash["A1"] = "Crypto Analytics Control Panel"
    ws_dash["A1"].font = font_title
    ws_dash["A2"] = f"Automated pipeline telemetry — Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_dash["A2"].font = font_subtitle

    # --- KPI cards ---
    kpi_configs = [
        ("A4", "DATASET LOG VOLUME", f"{len(df_updated)} Rows"),
        ("D4", "MONITORED TICKERS", f"{df_updated['symbol'].nunique()} Unique Coins"),
    ]
    for target_cell, label, val in kpi_configs:
        start_col, start_row = target_cell[0], int(target_cell[1])
        end_col = chr(ord(start_col) + 1)

        ws_dash.merge_cells(f"{start_col}{start_row}:{end_col}{start_row}")
        ws_dash[f"{start_col}{start_row}"] = label
        ws_dash[f"{start_col}{start_row}"].font = font_kpi_label
        ws_dash[f"{start_col}{start_row}"].alignment = align_center

        ws_dash.merge_cells(f"{start_col}{start_row+1}:{end_col}{start_row+1}")
        ws_dash[f"{start_col}{start_row+1}"] = val
        ws_dash[f"{start_col}{start_row+1}"].font = font_kpi_value
        ws_dash[f"{start_col}{start_row+1}"].alignment = align_center

        for r in range(start_row, start_row + 2):
            for c in range(ord(start_col) - 64, ord(end_col) - 63):
                ws_dash.cell(row=r, column=c).fill = fill_kpi
                ws_dash.cell(row=r, column=c).border = border_cell

    # --- Market leaders table ---
    ws_dash["A7"] = "Market Performance Leaders (Latest Snapshot Window)"
    ws_dash["A7"].font = font_section

    headers_dash = ["Insight Metric", "Asset Name", "Ticker", "Current Value", "24h Shift %"]
    for idx, h_text in enumerate(headers_dash, start=1):
        c = ws_dash.cell(row=8, column=idx, value=h_text)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_cell

    latest_time = df_updated["snapshot_time"].max()
    df_latest_snapshot = df_updated[df_updated["snapshot_time"] == latest_time]

    top_gainer = df_latest_snapshot.sort_values(by="price_change_percentage_24h", ascending=False).iloc[0]
    top_loser = df_latest_snapshot.sort_values(by="price_change_percentage_24h", ascending=True).iloc[0]

    leaders_rows = [
        ("Top Market Gainer", top_gainer["name"], top_gainer["symbol"], top_gainer["current_price"], top_gainer["price_change_percentage_24h"]),
        ("Top Market Loser", top_loser["name"], top_loser["symbol"], top_loser["current_price"], top_loser["price_change_percentage_24h"]),
    ]

    for r_idx, row_data in enumerate(leaders_rows, start=9):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            if c_idx == 1:
                cell.font = Font(name=FONT_NAME, size=11, bold=True)
            elif c_idx == 3:
                cell.alignment = align_center
            elif c_idx == 4:
                cell.number_format = "$#,##0.00"
                cell.alignment = align_right
            elif c_idx == 5:
                # price_change_percentage_24h is already normalized to a fraction in Step 1
                cell.number_format = "0.00%"
                cell.alignment = align_right
                cell.font = font_positive if val > 0 else font_negative

    # --- Price Trend Chart ---
    # Pivot the timeline into wide form (one column per coin) on a hidden
    # helper sheet, then plot it as a line chart embedded on the Dashboard.
    ws_dash["A14"] = "Price Trend (All Snapshots)"
    ws_dash["A14"].font = font_section

    if "ChartData" in wb.sheetnames:
        del wb["ChartData"]
    ws_chart_data = wb.create_sheet("ChartData")
    ws_chart_data.sheet_state = "hidden"

    df_pivot = (
        df_updated
        .pivot_table(index="snapshot_time", columns="symbol", values="current_price", aggfunc="last")
        .sort_index()
        .reset_index()
    )

    for r in dataframe_to_rows(df_pivot, index=False, header=True):
        ws_chart_data.append(r)

    if len(df_pivot) >= 1 and len(df_pivot.columns) >= 2:
        chart = LineChart()
        chart.title = "Price Trend by Coin"
        chart.y_axis.title = "Price (USD)"
        chart.x_axis.title = "Snapshot"
        chart.height = 9
        chart.width = 20

        max_row = ws_chart_data.max_row
        max_col = ws_chart_data.max_column

        data = Reference(ws_chart_data, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
        categories = Reference(ws_chart_data, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws_dash.add_chart(chart, "A15")

    # --- Timeline sheet formatting ---
    ws_timeline.row_dimensions[1].height = 26
    ws_timeline.freeze_panes = "A2"

    for cell in ws_timeline[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    for row in ws_timeline.iter_rows(min_row=2, max_row=ws_timeline.max_row, min_col=1, max_col=ws_timeline.max_column):
        for cell in row:
            cell.font = font_data
            cell.border = border_cell
            col_name = ws_timeline.cell(row=1, column=cell.column).value

            if col_name == "current_price":
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif col_name in ["market_cap", "total_volume"]:
                cell.alignment = align_right
                cell.number_format = "$#,##0"
            # price_change_percentage_24h is already normalized to a fraction in Step 1
            elif col_name == "price_change_percentage_24h":
                cell.alignment = align_right
                cell.number_format = "0.00%"
                if cell.value is not None:
                    cell.font = font_positive if cell.value > 0 else font_negative
            elif col_name in ["symbol", "snapshot_time"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # --- Column widths ---
    financial_cols = {"C": 16, "D": 22, "E": 22}
    for ws in [ws_dash, ws_timeline]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if ws == ws_timeline and col_letter in financial_cols:
                ws.column_dimensions[col_letter].width = financial_cols[col_letter]
            else:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save(output_file)
    logger.info("Excel report saved: '%s'", output_file)


def is_snapshot_too_soon(output_file: str, min_interval_minutes: int) -> bool:
    """True if the last saved snapshot is more recent than the configured interval."""
    if min_interval_minutes <= 0 or not os.path.exists(output_file):
        return False

    df_historical = pd.read_excel(output_file, sheet_name="Crypto Market Timeline")
    if df_historical.empty:
        return False

    last_time = pd.to_datetime(df_historical["snapshot_time"]).max()
    elapsed_minutes = (datetime.datetime.now() - last_time).total_seconds() / 60
    if elapsed_minutes < min_interval_minutes:
        logger.info(
            "Skipping: last snapshot was %.1f min ago, below MIN_SNAPSHOT_INTERVAL_MINUTES=%s.",
            elapsed_minutes, min_interval_minutes,
        )
        return True
    return False


# =============================================================================
# Orchestration
# =============================================================================
def main() -> int:
    setup_logging()
    logger.info("=== Crypto ETL Pipeline starting ===")

    if is_snapshot_too_soon(config.OUTPUT_FILE, config.MIN_SNAPSHOT_INTERVAL_MINUTES):
        return 0

    raw_json = fetch_crypto_data_with_retry()
    if not raw_json:
        logger.error("Pipeline failure: unable to fetch data. Stopping before any file writes.")
        return 1

    df_clean = clean_raw_data(raw_json)

    try:
        validate_data(df_clean)
    except ValueError as e:
        logger.error("Data validation failed: %s", e)
        return 1

    backup_existing_file(config.OUTPUT_FILE, config.BACKUP_DIR, config.MAX_BACKUPS)
    df_updated = load_and_merge_history(config.OUTPUT_FILE, df_clean)
    build_excel_report(df_updated, config.OUTPUT_FILE)

    logger.info("=== Pipeline completed successfully ===")
    return 0


if __name__ == "__main__":
    if sys.stdin.isatty() and "--no-prompt" not in sys.argv:
        config.COIN_IDS = prompt_for_coins(config.COIN_IDS)
    sys.exit(main())
